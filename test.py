import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from tabulate import tabulate
import numpy as np
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill

file_path = "Assignment [SST Admissions Team].xlsx"
df = pd.read_excel(file_path, sheet_name="Raw Data")

df.columns = df.columns.str.strip()
print("\n📌 Column Names in Dataset:")
print(tabulate(pd.DataFrame(df.columns, columns=["Column Names"]), headers="keys", tablefmt="grid"))

MIN_X_SCORE = 85
MIN_XII_SCORE = 80

def normalize_grade(grade):
    grade_to_percentage = {
        "A": 95,
        "B+": 85,
        "B": 75,
        "C+": 65,
        "C": 55,
        "D": 45,
        "F": 35
    }
    return grade_to_percentage.get(grade, np.nan)

def normalize_cgpa(cgpa):
    return cgpa * 9.5

df["x score"] = df.apply(
    lambda row: normalize_grade(row["x score"]) if row["x grade type"] == "Grade"
    else normalize_cgpa(row["x score"]) if row["x grade type"] == "CGPA"
    else row["x score"], axis=1
)

df["xii overall score"] = df.apply(
    lambda row: normalize_grade(row["xii overall score"]) if row["xii score type"] == "Grade"
    else normalize_cgpa(row["xii overall score"]) if row["xii score type"] == "CGPA"
    else row["xii overall score"], axis=1
)

df["x score"].fillna(df["x score"].median(), inplace=True)
df["xii overall score"].fillna(df["xii overall score"].median(), inplace=True)

shortlisted_df = df[(df["x score"] >= MIN_X_SCORE) & (df["xii overall score"] >= MIN_XII_SCORE)]

def assign_tier(row):
    if row["x score"] >= 90 and row["xii overall score"] >= 90:
        return "Tier 1"
    elif row["x score"] >= 85 and row["xii overall score"] >= 85:
        return "Tier 2"
    else:
        return "Tier 3"

shortlisted_df["Tier"] = shortlisted_df.apply(assign_tier, axis=1)

shortlisted_df = shortlisted_df.sort_values(by=["Tier", "xii overall score"], ascending=[True, False])

print("\n✅ Shortlisted Students:")
print(tabulate(shortlisted_df[["User ID", "first name", "x score", "xii overall score", "Tier"]].fillna("N/A"), 
               headers="keys", tablefmt="plain", stralign="center", numalign="center", 
               colalign=("center", "center", "center", "center", "center")))

shortlisted_file = "Shortlisted_Students.xlsx"
shortlisted_df.to_excel(shortlisted_file, index=False)

def format_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    wb.save(file_path)

format_excel(shortlisted_file)
print(f"\n✅ Shortlisted data saved to '{shortlisted_file}'")

total_applicants = len(df)
total_shortlisted = len(shortlisted_df)
avg_x_score = shortlisted_df["x score"].mean()
avg_overall_score = shortlisted_df["xii overall score"].mean()
top_students = shortlisted_df.nlargest(3, "xii overall score")

top_students = top_students.copy()
top_students["math score"] = top_students["math score"].astype(str).str.replace(",", " ")

summary_data = {
    "Metric": ["Total Applicants", "Total Shortlisted", "Average X Score", "Average Overall Score"],
    "Value": [total_applicants, total_shortlisted, round(avg_x_score, 2), round(avg_overall_score, 2)]
}
print("\n📌 Summary Statistics:")
print(tabulate(pd.DataFrame(summary_data), headers="keys", tablefmt="plain"))

columns_to_display = ["User ID", "first name", "gender", "10th board", "x score", "12th board", 
                      "xii overall score", "math score", "family income", "Tier"]
top_students_cleaned = top_students[columns_to_display].fillna("N/A")

top_students_cleaned["first name"] = top_students_cleaned["first name"].str[:10]
top_students_cleaned["family income"] = top_students_cleaned["family income"].str.replace("LPA", "L")

print("\n🏆 Top 3 Students:")
print(tabulate(top_students_cleaned, headers="keys", tablefmt="plain", 
               stralign="center", numalign="center", 
               colalign=("center", "center", "center", "center", "center", 
                         "center", "center", "center", "center", "center")))

top_students_cleaned.to_excel("Top_Students_Without_JEE.xlsx", index=False)
format_excel("Top_Students_Without_JEE.xlsx")
print("\n✅ Top 3 students (without JEE Mains score) saved to 'Top_Students_Without_JEE.xlsx'")

plt.figure(figsize=(8,5))
sns.histplot(shortlisted_df["x score"], bins=10, kde=True, color="blue")
plt.title("Distribution of X Scores")
plt.xlabel("X Score")
plt.ylabel("Frequency")
plt.show()

plt.figure(figsize=(8,5))
sns.histplot(shortlisted_df["xii overall score"], bins=10, kde=True, color="green")
plt.title("Distribution of Overall Scores")
plt.xlabel("Overall Score")
plt.ylabel("Frequency")
plt.show()

top_5_students = shortlisted_df.nlargest(5, "xii overall score")

plt.figure(figsize=(8,5))
sns.barplot(x="first name", y="xii overall score", data=top_5_students, hue="first name", palette="magma")
plt.title("Top 5 Students by Overall Score")
plt.xlabel("Student Name")
plt.ylabel("Overall Score")
plt.xticks(rotation=45)
plt.legend().remove()
plt.show()

plt.figure(figsize=(8,5))
sns.barplot(x="first name", y="x score", data=top_5_students, hue="first name", palette="coolwarm")
plt.title("Top 5 Students by X Score")
plt.xlabel("Student Name")
plt.ylabel("X Score")
plt.xticks(rotation=45)
plt.legend().remove()
plt.show()

INTERVIEWERS = ["Interviewer 1", "Interviewer 2", "Interviewer 3"]
SLOT_DURATION = 30
SLOTS_PER_DAY = (3 * 60) // SLOT_DURATION
NUM_ROUNDS = 2

shortlisted_df = shortlisted_df.sample(frac=1, random_state=42).reset_index(drop=True)
interview_schedule = []

for round_num in range(1, NUM_ROUNDS + 1):
    for i, student in shortlisted_df.iterrows():
        interviewer = INTERVIEWERS[i % len(INTERVIEWERS)]
        slot = (i % SLOTS_PER_DAY) + 1
        interview_schedule.append([student["User ID"], student["first name"], round_num, interviewer, f"Slot {slot}"])

interview_df = pd.DataFrame(interview_schedule, columns=["User ID", "first name", "Round", "Interviewer", "Slot"])
print("\n✅ Interview Schedule:")
print(tabulate(interview_df, headers="keys", tablefmt="plain"))

interview_schedule_file = "Interview_Schedule.xlsx"
interview_df.to_excel(interview_schedule_file, index=False)
format_excel(interview_schedule_file)
print(f"\n✅ Interview schedule saved to '{interview_schedule_file}'")
